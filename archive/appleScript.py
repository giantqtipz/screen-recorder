import os
from openpyxl import Workbook, load_workbook
import webbrowser

wb = load_workbook('gymshark_influencers.xlsx')
ws = wb.active

urls = ws['H']
tiktok = ws['F']

handle = ''

command = """ osascript -e '

--  # Setup to do a screen recording.

tell application "QuickTime Player" to new screen recording

--  # Start the screen recording.

tell application "System Events" to tell process "Screen Shot"
    repeat until exists button "Record" of its front window
        delay 0.1
    end repeat
    click button "Record" of its front window
end tell

--  # Set the time in seconds you want the recording to be.

delay 10

--  # Stop the recording.

tell application "System Events" to ¬
    click menu bar item 1 ¬
        of menu bar 1 ¬
        of application process "screencaptureui"

tell application "QuickTime Player" to quit

'"""

for i, cell in enumerate(urls):
    url = cell.value
    handle = tiktok[i].value
    webbrowser.open(url, new=0)
    os.system(command)