from openpyxl import Workbook, load_workbook
import os
import webbrowser

# from webdriver_manager.chrome import ChromeDriverManager
# from selenium import webdriver

wb = load_workbook('./excel/gymshark_influencers.xlsx')
ws = wb.active

urls = ws['H']
tiktok = ws['F']

def record(handle):
    os.system(f"""ffmpeg -f avfoundation -i "1:0" -t 00:00:23 -y -r 10 recordings/{handle}.mov""")

for i, cell in enumerate(urls):
    url = cell.value
    handle = tiktok[i].value
    webbrowser.open(url)
    record(handle)


# ----------  Selenium Attempt  ----------

# driver = webdriver.Chrome(ChromeDriverManager().install())
# url = 'https://www.tiktok.com/@flexforall/video/7062901737237220654?is_copy_url=1&is_from_webapp=v1&lang=en'

# driver = webdriver.Chrome()
# driver.get(url)

# video = driver.find_element_by_xpath("//li[@data-e2e='video-author-uniqueid']")
# print(video)

# ----------  Selenium Attempt  ----------