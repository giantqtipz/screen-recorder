from openpyxl import load_workbook

def open_workbook(filename, url_column):

    wb = load_workbook(filename)
    ws = wb.active

    urls = ws[url_column]
    return urls

